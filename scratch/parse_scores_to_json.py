import openpyxl
import json

wb = openpyxl.load_workbook('data/2568_RM&IC_OFI Improvement Plan_update_20260622_154253.xlsx', data_only=True)
sheet = wb.worksheets[1] # 2nd sheet

scores = []
for r in range(3, 30):
    code = "RM&IC"
    name = sheet.cell(r, 4).value
    if not name:
        continue
    
    name = name.strip()
    if name == "คู่มือ" or name == "คะแนน" or name == "รวมคะแนน" or name == "∑" or "รวม" in name or name == "":
        continue
        
    y63 = sheet.cell(r, 9).value or 0
    y64 = sheet.cell(r, 10).value or 0
    y65 = sheet.cell(r, 11).value or 0
    y66 = sheet.cell(r, 12).value or 0
    y67 = sheet.cell(r, 13).value or 0
    y68 = sheet.cell(r, 14).value or 0
    
    try:
        y63 = float(y63)
    except:
        y63 = 0.0
    try:
        y64 = float(y64)
    except:
        y64 = 0.0
    try:
        y65 = float(y65)
    except:
        y65 = 0.0
    try:
        y66 = float(y66)
    except:
        y66 = 0.0
    try:
        y67 = float(y67)
    except:
        y67 = 0.0
    try:
        y68 = float(y68) if y68 is not None and y68 != "" else 0.0
    except:
        y68 = 0.0

    scores.append({
        "code": code,
        "name": name,
        "y63": round(y63, 4),
        "y64": round(y64, 4),
        "y65": round(y65, 4),
        "y66": round(y66, 4),
        "y67": round(y67, 4),
        "y68": round(y68, 4)
    })

# Save to a json file to check if it's correct
with open('scratch/scores_output.json', 'w', encoding='utf-8') as f:
    json.dump(scores, f, indent=4, ensure_ascii=False)
print("SUCCESS")
